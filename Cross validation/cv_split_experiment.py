"""
cv_split_experiment.py
----------------------
Experiment: vary train/test split ratios and compare OOS performance.
Evaluates robustness of Baseline and Augmented models across multiple splits.

Usage:
    python cv_split_experiment.py

Input:  Compare_y_to_market.xlsx  (merged_data sheet)
Output: CV_Split_Results.xlsx
"""

import warnings
warnings.filterwarnings("ignore")

import pandas as pd
import numpy as np
from sklearn.linear_model import Ridge
from sklearn.preprocessing import StandardScaler
from pathlib import Path

# ── Config ────────────────────────────────────────────────────────────────────
INPUT_FILE  = "Compare_y_to_market.xlsx"
OUTPUT_FILE = "CV_Split_Results.xlsx"

TRAIN_RATIOS = [0.50, 0.55, 0.60, 0.65, 0.70, 0.75, 0.80]

BASELINE_FEATURES = [
    "surprise", "goal_diff", "venue_home",
    "importance_0_1", "opponent_strength_0_1",
    "recent_return_5d", "recent_volatility_20d",
    "market_index_return_1d", "liquidity_thinness_0_1",
]

PERSONA_FEATURES = [
    "mean_net_demand", "buy_share", "sell_share",
    "liq_weighted_pressure",
    "mean_net_demand__investor_fan",
    "buy_share__investor_fan", "sell_share__investor_fan",
    "mean_net_demand__fanatical_fan",
    "sell_share__fanatical_fan",
    "mean_net_demand__temporary_fan",
    "mean_net_demand__local_fan",
]

AUGMENTED_FEATURES = BASELINE_FEATURES + PERSONA_FEATURES

TARGET = "y_realized"

ALPHA = 1.0  # Ridge regularization


# ── Helpers ───────────────────────────────────────────────────────────────────

def _safe_features(df: pd.DataFrame, feats: list) -> list:
    """Return only features that exist in df."""
    return [f for f in feats if f in df.columns]


def _metrics(y_true: np.ndarray, y_hat: np.ndarray) -> dict:
    """Compute regression + direction metrics."""
    n         = len(y_true)
    residuals = y_true - y_hat
    ss_res    = np.sum(residuals ** 2)
    ss_tot    = np.sum((y_true - y_true.mean()) ** 2)
    r2_oos    = 1 - ss_res / ss_tot if ss_tot > 0 else np.nan
    rmse      = np.sqrt(np.mean(residuals ** 2))
    mae       = np.mean(np.abs(residuals))
    hit_rate  = np.mean(np.sign(y_true) == np.sign(y_hat))
    corr      = np.corrcoef(y_true, y_hat)[0, 1] if n > 2 else np.nan
    return dict(n=n, r2_oos=r2_oos, rmse=rmse, mae=mae,
                hit_rate=hit_rate, corr=corr)


def run_rolling_oos(df: pd.DataFrame, features: list,
                    min_train: int) -> dict:
    """
    Rolling-origin expanding window OOS regression.
    Returns aggregated metrics over all OOS predictions.
    """
    feats = _safe_features(df, features)
    df_   = df.dropna(subset=feats + [TARGET]).reset_index(drop=True)
    n     = len(df_)

    y_true_all, y_hat_all = [], []

    for t in range(min_train, n):
        train = df_.iloc[:t]
        test  = df_.iloc[[t]]

        X_train = train[feats].values
        y_train = train[TARGET].values
        X_test  = test[feats].values
        y_test  = test[TARGET].values[0]

        scaler  = StandardScaler()
        X_tr_sc = scaler.fit_transform(X_train)
        X_te_sc = scaler.transform(X_test)

        model   = Ridge(alpha=ALPHA)
        model.fit(X_tr_sc, y_train)
        y_pred  = model.predict(X_te_sc)[0]

        y_true_all.append(y_test)
        y_hat_all.append(y_pred)

    return _metrics(np.array(y_true_all), np.array(y_hat_all))


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    # Load data
    src = Path(INPUT_FILE)
    if not src.exists():
        src = Path("../uploads") / INPUT_FILE
    if not src.exists():
        raise FileNotFoundError(f"Cannot find {INPUT_FILE}")

    df = pd.read_excel(src, sheet_name="merged_data")
    df = df.sort_values("match_date").reset_index(drop=True)
    total = len(df)
    print(f"Loaded {total} matches from {src.name}")

    # Run experiments
    rows = []
    for ratio in TRAIN_RATIOS:
        min_train = max(20, int(total * ratio))
        oos_n     = total - min_train
        pct_test  = round(1 - ratio, 2)

        print(f"\nTrain {int(ratio*100)}% / Test {int(pct_test*100)}%"
              f"  (min_train={min_train}, OOS={oos_n})")

        for model_name, feats in [("Baseline", BASELINE_FEATURES),
                                   ("Augmented", AUGMENTED_FEATURES)]:
            m = run_rolling_oos(df, feats, min_train)
            rows.append({
                "train_pct"  : f"{int(ratio*100)}%",
                "test_pct"   : f"{int(pct_test*100)}%",
                "min_train"  : min_train,
                "oos_n"      : m["n"],
                "model"      : model_name,
                "r2_oos"     : round(m["r2_oos"], 4),
                "hit_rate"   : round(m["hit_rate"], 4),
                "rmse"       : round(m["rmse"], 5),
                "mae"        : round(m["mae"], 5),
                "corr"       : round(m["corr"], 4),
            })
            print(f"  {model_name:10s}  r2_oos={m['r2_oos']:+.4f}"
                  f"  hit_rate={m['hit_rate']:.1%}"
                  f"  corr={m['corr']:+.3f}")

    results = pd.DataFrame(rows)

    # ── Build Excel output ────────────────────────────────────────────────────
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "CV Results"

    # Colour palette
    C_HEADER  = "1E3A5F"   # dark navy
    C_BASE    = "D6E4F0"   # light blue — baseline rows
    C_AUG     = "D6F0DC"   # light green — augmented rows
    C_SEP     = "F2F2F2"   # gray — ratio separator
    C_BEST_POS = "2ECC71"  # green — best positive r2
    C_BEST_DIR = "3498DB"  # blue — best hit rate

    thin = Side(style="thin", color="AAAAAA")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    ws["A1"] = "Cross-Validation Split Ratio Experiment — MANU Simulation"
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor=C_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:J2")
    ws["A2"] = (f"Total matches: {total}  |  Ridge α={ALPHA}  |"
                f"  Rolling-origin expanding window")
    ws["A2"].font      = Font(name="Arial", size=10, italic=True, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Column headers ────────────────────────────────────────────────────────
    headers = ["Train %", "Test %", "Min Train", "OOS n",
               "Model", "R² OOS", "Hit Rate", "RMSE", "MAE", "Corr"]
    col_widths = [10, 9, 11, 9, 13, 11, 11, 11, 11, 10]

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor=C_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = bdr
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────────
    row_idx = 4
    prev_ratio = None

    for _, rec in results.iterrows():
        is_new_ratio = rec["train_pct"] != prev_ratio
        if is_new_ratio and prev_ratio is not None:
            # blank separator row
            for ci in range(1, 11):
                ws.cell(row=row_idx, column=ci).fill = PatternFill("solid", fgColor=C_SEP)
            row_idx += 1
        prev_ratio = rec["train_pct"]

        is_aug   = rec["model"] == "Augmented"
        row_fill = PatternFill("solid", fgColor=C_AUG if is_aug else C_BASE)

        vals = [rec["train_pct"], rec["test_pct"], rec["min_train"],
                rec["oos_n"],     rec["model"],
                rec["r2_oos"],    rec["hit_rate"],
                rec["rmse"],      rec["mae"],      rec["corr"]]

        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=ci, value=v)
            cell.fill      = row_fill
            cell.border    = bdr
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center")

        # Number formats
        ws.cell(row=row_idx, column=6).number_format = "+0.0000;-0.0000;0.0000"
        ws.cell(row=row_idx, column=7).number_format = "0.0%"
        ws.cell(row=row_idx, column=8).number_format = "0.00000"
        ws.cell(row=row_idx, column=9).number_format = "0.00000"
        ws.cell(row=row_idx, column=10).number_format = "+0.000;-0.000;0.000"

        row_idx += 1

    # ── Summary pivot: Baseline vs Augmented hit-rate difference ─────────────
    row_idx += 2
    ws.cell(row=row_idx, column=1,
            value="Summary: Hit Rate Difference (Augmented − Baseline)").font = \
        Font(name="Arial", bold=True, size=11)
    row_idx += 1

    pivot_headers = ["Train %", "Test %", "OOS n",
                     "Baseline Hit Rate", "Augmented Hit Rate",
                     "Δ Hit Rate", "Baseline R² OOS", "Augmented R² OOS"]
    for ci, h in enumerate(pivot_headers, start=1):
        cell = ws.cell(row=row_idx, column=ci, value=h)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor=C_HEADER)
        cell.alignment = Alignment(horizontal="center")
        cell.border    = bdr
    row_idx += 1

    base_df = results[results["model"] == "Baseline"].reset_index(drop=True)
    aug_df  = results[results["model"] == "Augmented"].reset_index(drop=True)

    for i in range(len(base_df)):
        delta_hit = aug_df.loc[i, "hit_rate"] - base_df.loc[i, "hit_rate"]
        vals = [
            base_df.loc[i, "train_pct"],
            base_df.loc[i, "test_pct"],
            base_df.loc[i, "oos_n"],
            base_df.loc[i, "hit_rate"],
            aug_df.loc[i,  "hit_rate"],
            delta_hit,
            base_df.loc[i, "r2_oos"],
            aug_df.loc[i,  "r2_oos"],
        ]
        # Highlight row: green if augmented dominates on hit rate
        row_color = "E8F8F0" if delta_hit > 0 else "FDF2F2"
        rf = PatternFill("solid", fgColor=row_color)

        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=ci, value=v)
            cell.fill      = rf
            cell.border    = bdr
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center")

        ws.cell(row=row_idx, column=4).number_format = "0.0%"
        ws.cell(row=row_idx, column=5).number_format = "0.0%"
        ws.cell(row=row_idx, column=6).number_format = "+0.0%;-0.0%;0.0%"
        ws.cell(row=row_idx, column=7).number_format = "+0.0000;-0.0000;0.0000"
        ws.cell(row=row_idx, column=8).number_format = "+0.0000;-0.0000;0.0000"
        row_idx += 1

    # ── Legend ────────────────────────────────────────────────────────────────
    row_idx += 1
    legends = [
        ("■", C_BASE,  "Baseline model (match + market features only)"),
        ("■", C_AUG,   "Augmented model (+ LLM persona features)"),
        ("■", "E8F8F0","Augmented hit rate > Baseline (LLM adds directional signal)"),
        ("■", "FDF2F2","Augmented hit rate ≤ Baseline"),
    ]
    for sym, color, desc in legends:
        ws.cell(row=row_idx, column=1, value=sym).font = \
            Font(name="Arial", color=color, size=14)
        ws.cell(row=row_idx, column=2, value=desc).font = \
            Font(name="Arial", size=9, color="444444")
        ws.merge_cells(f"B{row_idx}:J{row_idx}")
        row_idx += 1

    # ── Freeze panes ─────────────────────────────────────────────────────────
    ws.freeze_panes = "A4"

    out_path = Path(OUTPUT_FILE)
    wb.save(out_path)
    print(f"\nSaved → {out_path.resolve()}")
    return str(out_path)


if __name__ == "__main__":
    main()
